"""The cross-reference as a spreadsheet, one row per ISO control."""
from __future__ import annotations

from typing import Any, List

from .model import COVERAGE_TEXT, Dataset

FILL = {"full": "C6EFCE", "partial": "FFEB9C", "none": "F2F2F2"}


def write(path: str, dataset: Dataset, lang: str = "es") -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depends on environment
        raise RuntimeError("openpyxl is required for the spreadsheet: pip install openpyxl") from exc

    es = lang == "es"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Equivalencias" if es else "Cross-reference"

    headers = (["Control ISO", "Titulo ISO", "Tema", "ENS", "Cobertura ENS", "NIS2",
                "Cobertura NIS2", "DORA", "Cobertura DORA", "Fuentes", "Estado"] if es else
               ["ISO control", "ISO title", "Theme", "ENS", "ENS coverage", "NIS2",
                "NIS2 coverage", "DORA", "DORA coverage", "Sources", "Status"])
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    row_index = 2
    for control in dataset.all_controls("ISO"):
        links = dataset.forward.get(control.id, {})
        values: List[Any] = [control.id, control.name(lang), control.family_title.get(lang, "")]
        sources = set()
        for framework in ("ENS", "NIS2", "DORA"):
            targets = links.get(framework, [])
            values.append(", ".join(t.target for t in targets) or "—")
            coverage = dataset.coverage.get(control.id, {}).get(framework, "none")
            values.append(COVERAGE_TEXT[coverage][lang])
            sources.update(t.source for t in targets)
        values.append(", ".join(sorted(sources)) or "—")
        values.append(next(iter({l.status for l in
                                 [x for group in links.values() for x in group]}), "proposed"))
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=column in (2, 3, 4, 6, 8, 10))
        for offset, framework in ((5, "ENS"), (7, "NIS2"), (9, "DORA")):
            coverage = dataset.coverage.get(control.id, {}).get(framework, "none")
            sheet.cell(row=row_index, column=offset).fill = PatternFill("solid", fgColor=FILL[coverage])
        row_index += 1

    for index, width in enumerate([12, 46, 22, 26, 14, 26, 14, 26, 14, 26, 12], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_index - 1}"
    mark = sheet.cell(row=row_index + 1, column=1,
                      value="crossmap · mr7security · github.com/mr7security/crossmap")
    mark.font = Font(size=9, color="808080")

    # Second sheet: what each regime asks for that ISO does not cover.
    gaps = workbook.create_sheet("Huecos" if es else "Gaps")
    gap_headers = (["Marco", "Referencia", "Titulo", "Familia"] if es else
                   ["Framework", "Reference", "Title", "Family"])
    for column, header in enumerate(gap_headers, start=1):
        cell = gaps.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="7B2D26")
    from .query import orphans
    index = 2
    for framework in ("ENS", "NIS2", "DORA"):
        for control in orphans(dataset, framework):
            for column, value in enumerate(
                [framework, control.id, control.name(lang), control.family_title.get(lang, "")],
                start=1,
            ):
                gaps.cell(row=index, column=column, value=value).alignment = Alignment(
                    vertical="top", wrap_text=column in (3, 4))
            index += 1
    for column, width in enumerate([12, 18, 60, 40], start=1):
        gaps.column_dimensions[get_column_letter(column)].width = width
    gaps.freeze_panes = "A2"

    workbook.save(path)
    return path
